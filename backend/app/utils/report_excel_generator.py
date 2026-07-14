from collections.abc import Iterable, Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


class ReportExcelGenerator:
    """
    Generador reutilizable de reportes profesionales en formato Excel.

    No consulta la base de datos ni contiene lógica de negocio.
    Recibe únicamente información previamente procesada.
    """

    TITLE_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    SUMMARY_LABEL_FILL = PatternFill(
        fill_type="solid",
        fgColor="EAF2F8",
    )

    WHITE_FONT = Font(
        color="FFFFFF",
        bold=True,
    )

    TITLE_FONT = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )

    HEADER_FONT = Font(
        color="1F1F1F",
        bold=True,
        size=10,
    )

    BODY_FONT = Font(
        color="1F1F1F",
        size=10,
    )

    SUMMARY_LABEL_FONT = Font(
        color="1F1F1F",
        bold=True,
        size=10,
    )

    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    CURRENCY_FORMAT = '"S/" #,##0.00'
    DECIMAL_FORMAT = '#,##0.000'
    INTEGER_FORMAT = '#,##0'
    PERCENTAGE_FORMAT = '0.00" %"'
    DATE_FORMAT = 'dd/mm/yyyy'
    DATETIME_FORMAT = 'dd/mm/yyyy hh:mm'

    MIN_COLUMN_WIDTH = 10
    MAX_COLUMN_WIDTH = 45

    @classmethod
    def generate(
        cls,
        *,
        file_path: Path | str,
        report_title: str,
        sheet_name: str,
        columns: Sequence[Mapping[str, Any]],
        rows: Iterable[Mapping[str, Any]],
        summary: Mapping[str, Any] | None = None,
        metadata: Mapping[str, Any] | None = None,
    ) -> Path:
        """
        Genera un archivo Excel profesional.

        Cada columna debe tener como mínimo:

        {
            "key": "total_amount",
            "label": "Total"
        }

        Opcionalmente puede incluir:

        {
            "format": "currency",
            "width": 18
        }
        """

        destination = Path(file_path).expanduser().resolve()

        cls._validate_input(
            destination=destination,
            report_title=report_title,
            sheet_name=sheet_name,
            columns=columns,
        )

        normalized_rows = list(rows)

        destination.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = cls._sanitize_sheet_name(sheet_name)

        current_row = 1

        current_row = cls._write_title(
            worksheet=worksheet,
            row_number=current_row,
            report_title=report_title,
            column_count=len(columns),
        )

        if metadata:
            current_row = cls._write_metadata(
                worksheet=worksheet,
                row_number=current_row,
                metadata=metadata,
                column_count=len(columns),
            )

        if summary:
            current_row = cls._write_summary(
                worksheet=worksheet,
                row_number=current_row,
                summary=summary,
                column_count=len(columns),
            )

        header_row = current_row

        cls._write_headers(
            worksheet=worksheet,
            row_number=header_row,
            columns=columns,
        )

        data_start_row = header_row + 1

        data_end_row = cls._write_rows(
            worksheet=worksheet,
            start_row=data_start_row,
            columns=columns,
            rows=normalized_rows,
        )

        cls._configure_table(
            worksheet=worksheet,
            header_row=header_row,
            data_end_row=data_end_row,
            column_count=len(columns),
        )

        cls._configure_worksheet(
            worksheet=worksheet,
            columns=columns,
            header_row=header_row,
            data_start_row=data_start_row,
            data_end_row=data_end_row,
        )

        workbook.save(destination)

        if not destination.exists():
            raise RuntimeError(
                "El archivo Excel no pudo ser generado."
            )

        return destination

    @classmethod
    def _validate_input(
        cls,
        *,
        destination: Path,
        report_title: str,
        sheet_name: str,
        columns: Sequence[Mapping[str, Any]],
    ) -> None:
        if destination.suffix.lower() != ".xlsx":
            raise ValueError(
                "La ruta del reporte Excel debe terminar en .xlsx."
            )

        if not report_title or not report_title.strip():
            raise ValueError(
                "El título del reporte es obligatorio."
            )

        if not sheet_name or not sheet_name.strip():
            raise ValueError(
                "El nombre de la hoja es obligatorio."
            )

        if not columns:
            raise ValueError(
                "El reporte debe contener al menos una columna."
            )

        column_keys: set[str] = set()

        for index, column in enumerate(columns, start=1):
            key = str(column.get("key", "")).strip()
            label = str(column.get("label", "")).strip()

            if not key:
                raise ValueError(
                    f"La columna {index} no tiene una clave válida."
                )

            if not label:
                raise ValueError(
                    f"La columna {index} no tiene una etiqueta válida."
                )

            if key in column_keys:
                raise ValueError(
                    f"La clave de columna está duplicada: {key}"
                )

            column_keys.add(key)

    @classmethod
    def _write_title(
        cls,
        *,
        worksheet: Worksheet,
        row_number: int,
        report_title: str,
        column_count: int,
    ) -> int:
        worksheet.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=column_count,
        )

        title_cell = worksheet.cell(
            row=row_number,
            column=1,
            value=report_title.strip(),
        )

        title_cell.fill = cls.TITLE_FILL
        title_cell.font = cls.TITLE_FONT
        title_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        worksheet.row_dimensions[row_number].height = 28

        return row_number + 2

    @classmethod
    def _write_metadata(
        cls,
        *,
        worksheet: Worksheet,
        row_number: int,
        metadata: Mapping[str, Any],
        column_count: int,
    ) -> int:
        for label, value in metadata.items():
            worksheet.cell(
                row=row_number,
                column=1,
                value=str(label),
            )

            value_cell = worksheet.cell(
                row=row_number,
                column=2,
                value=cls._normalize_cell_value(value),
            )

            worksheet.cell(
                row=row_number,
                column=1,
            ).font = cls.SUMMARY_LABEL_FONT

            cls._apply_automatic_number_format(
                cell=value_cell,
                value=value,
            )

            if column_count >= 2:
                worksheet.merge_cells(
                    start_row=row_number,
                    start_column=2,
                    end_row=row_number,
                    end_column=column_count,
                )

            row_number += 1

        return row_number + 1

    @classmethod
    def _write_summary(
        cls,
        *,
        worksheet: Worksheet,
        row_number: int,
        summary: Mapping[str, Any],
        column_count: int,
    ) -> int:
        summary_items = list(summary.items())

        if not summary_items:
            return row_number

        available_pairs = max(1, column_count // 2)

        for item_index, (label, value) in enumerate(summary_items):
            pair_position = item_index % available_pairs

            if item_index > 0 and pair_position == 0:
                row_number += 1

            label_column = (pair_position * 2) + 1
            value_column = label_column + 1

            label_cell = worksheet.cell(
                row=row_number,
                column=label_column,
                value=str(label),
            )

            value_cell = worksheet.cell(
                row=row_number,
                column=value_column,
                value=cls._normalize_cell_value(value),
            )

            label_cell.fill = cls.SUMMARY_LABEL_FILL
            label_cell.font = cls.SUMMARY_LABEL_FONT
            label_cell.border = cls.THIN_BORDER
            label_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
            )

            value_cell.border = cls.THIN_BORDER
            value_cell.font = cls.BODY_FONT
            value_cell.alignment = Alignment(
                horizontal="right",
                vertical="center",
            )

            cls._apply_automatic_number_format(
                cell=value_cell,
                value=value,
            )

        return row_number + 2

    @classmethod
    def _write_headers(
        cls,
        *,
        worksheet: Worksheet,
        row_number: int,
        columns: Sequence[Mapping[str, Any]],
    ) -> None:
        for column_index, column_definition in enumerate(
            columns,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_index,
                value=str(column_definition["label"]),
            )

            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.border = cls.THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[row_number].height = 24

    @classmethod
    def _write_rows(
        cls,
        *,
        worksheet: Worksheet,
        start_row: int,
        columns: Sequence[Mapping[str, Any]],
        rows: Sequence[Mapping[str, Any]],
    ) -> int:
        if not rows:
            worksheet.cell(
                row=start_row,
                column=1,
                value="Sin registros para los filtros seleccionados.",
            )

            worksheet.cell(
                row=start_row,
                column=1,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
            )

            return start_row

        current_row = start_row

        for row_data in rows:
            for column_index, column_definition in enumerate(
                columns,
                start=1,
            ):
                key = str(column_definition["key"])
                value = row_data.get(key)

                cell = worksheet.cell(
                    row=current_row,
                    column=column_index,
                    value=cls._normalize_cell_value(value),
                )

                cell.font = cls.BODY_FONT
                cell.border = cls.THIN_BORDER
                cell.alignment = cls._get_alignment(
                    value=value,
                    column_format=column_definition.get("format"),
                )

                cls._apply_column_number_format(
                    cell=cell,
                    value=value,
                    column_format=column_definition.get("format"),
                )

            current_row += 1

        return current_row - 1

    @classmethod
    def _configure_table(
        cls,
        *,
        worksheet: Worksheet,
        header_row: int,
        data_end_row: int,
        column_count: int,
    ) -> None:
        if data_end_row <= header_row:
            return

        table_reference = (
            f"A{header_row}:"
            f"{get_column_letter(column_count)}{data_end_row}"
        )

        table = Table(
            displayName="NovaPOSReportTable",
            ref=table_reference,
        )

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        worksheet.add_table(table)

    @classmethod
    def _configure_worksheet(
        cls,
        *,
        worksheet: Worksheet,
        columns: Sequence[Mapping[str, Any]],
        header_row: int,
        data_start_row: int,
        data_end_row: int,
    ) -> None:
        worksheet.freeze_panes = f"A{data_start_row}"
        worksheet.auto_filter.ref = (
            f"A{header_row}:"
            f"{get_column_letter(len(columns))}{data_end_row}"
        )

        worksheet.sheet_view.showGridLines = False

        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        worksheet.print_title_rows = f"{header_row}:{header_row}"

        worksheet.print_area = (
            f"A1:{get_column_letter(len(columns))}{data_end_row}"
        )

        cls._adjust_column_widths(
            worksheet=worksheet,
            columns=columns,
            data_end_row=data_end_row,
        )

    @classmethod
    def _adjust_column_widths(
        cls,
        *,
        worksheet: Worksheet,
        columns: Sequence[Mapping[str, Any]],
        data_end_row: int,
    ) -> None:
        for column_index, column_definition in enumerate(
            columns,
            start=1,
        ):
            configured_width = column_definition.get("width")

            if configured_width is not None:
                width = float(configured_width)
            else:
                width = cls.MIN_COLUMN_WIDTH

                for row_number in range(1, data_end_row + 1):
                    cell_value = worksheet.cell(
                        row=row_number,
                        column=column_index,
                    ).value

                    if cell_value is None:
                        continue

                    value_length = len(str(cell_value)) + 2
                    width = max(width, value_length)

            width = min(
                max(width, cls.MIN_COLUMN_WIDTH),
                cls.MAX_COLUMN_WIDTH,
            )

            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width

    @classmethod
    def _apply_column_number_format(
        cls,
        *,
        cell: Cell,
        value: Any,
        column_format: Any,
    ) -> None:
        normalized_format = (
            str(column_format).strip().lower()
            if column_format is not None
            else ""
        )

        formats = {
            "currency": cls.CURRENCY_FORMAT,
            "money": cls.CURRENCY_FORMAT,
            "decimal": cls.DECIMAL_FORMAT,
            "quantity": cls.DECIMAL_FORMAT,
            "integer": cls.INTEGER_FORMAT,
            "percentage": cls.PERCENTAGE_FORMAT,
            "percent": cls.PERCENTAGE_FORMAT,
            "date": cls.DATE_FORMAT,
            "datetime": cls.DATETIME_FORMAT,
        }

        selected_format = formats.get(normalized_format)

        if selected_format:
            cell.number_format = selected_format
            return

        cls._apply_automatic_number_format(
            cell=cell,
            value=value,
        )

    @classmethod
    def _apply_automatic_number_format(
        cls,
        *,
        cell: Cell,
        value: Any,
    ) -> None:
        if isinstance(value, datetime):
            cell.number_format = cls.DATETIME_FORMAT
        elif isinstance(value, date):
            cell.number_format = cls.DATE_FORMAT
        elif isinstance(value, Decimal):
            cell.number_format = cls.DECIMAL_FORMAT

    @staticmethod
    def _normalize_cell_value(value: Any) -> Any:
        if value is None:
            return ""

        if isinstance(value, Decimal):
            return float(value)

        if isinstance(value, bool):
            return "Sí" if value else "No"

        if isinstance(value, Enum):
            return value.value

        if isinstance(value, (str, int, float, date, datetime)):
            return value

        return str(value)

    @staticmethod
    def _get_alignment(
        *,
        value: Any,
        column_format: Any,
    ) -> Alignment:
        normalized_format = (
            str(column_format).strip().lower()
            if column_format is not None
            else ""
        )

        numeric_formats = {
            "currency",
            "money",
            "decimal",
            "quantity",
            "integer",
            "percentage",
            "percent",
        }

        if normalized_format in numeric_formats:
            horizontal = "right"
        elif normalized_format in {"date", "datetime"}:
            horizontal = "center"
        elif isinstance(value, (int, float, Decimal)):
            horizontal = "right"
        elif isinstance(value, (date, datetime)):
            horizontal = "center"
        else:
            horizontal = "left"

        return Alignment(
            horizontal=horizontal,
            vertical="center",
            wrap_text=True,
        )

    @staticmethod
    def _sanitize_sheet_name(sheet_name: str) -> str:
        invalid_characters = {"\\", "/", "*", "?", ":", "[", "]"}

        sanitized = "".join(
            character
            for character in sheet_name.strip()
            if character not in invalid_characters
        )

        if not sanitized:
            sanitized = "Reporte"

        return sanitized[:31]